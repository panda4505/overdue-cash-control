"""Tabular file parser for import previews."""

from __future__ import annotations

import csv
import datetime as dt
import io
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import chardet
import pandas as pd
from openpyxl import load_workbook


ENCODING_FALLBACKS = [
    "utf-8",
    "utf-8-sig",
    "windows-1252",
    "iso-8859-1",
    "iso-8859-15",
    "windows-1250",
    "iso-8859-2",
]
SUMMARY_MARKERS = [
    "total",
    "totale",
    "totaux",
    "somme",
    "sous-total",
    "celkem",
    "součet",
    "suma",
    "gesamt",
    "summe",
    "zusammen",
]
DATE_PATTERNS = {
    "dot": re.compile(r"^\d{1,2}\.\d{1,2}\.\d{4}$"),
    "slash": re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$"),
    "iso": re.compile(r"^\d{4}-\d{2}-\d{2}$"),
}
SPACE_GROUP_COMMA_RE = re.compile(r"^\s*-?\d{1,3}(?: \d{3})+(?:,\d+)?\s*$")
DOT_GROUP_COMMA_RE = re.compile(r"^\s*-?\d{1,3}(?:\.\d{3})+(?:,\d+)?\s*$")
COMMA_GROUP_DOT_RE = re.compile(r"^\s*-?\d{1,3}(?:,\d{3})+(?:\.\d+)?\s*$")
SIMPLE_COMMA_DECIMAL_RE = re.compile(r"^\s*-?\d+,\d+\s*$")
PLAIN_DOT_DECIMAL_RE = re.compile(r"^\s*-?\d+\.\d+\s*$")


@dataclass
class ParseResult:
    """Output of the file parser. Everything downstream needs."""

    success: bool
    filename: str

    encoding: str | None = None
    delimiter: str | None = None
    header_row_index: int | None = None

    headers: list[str] = field(default_factory=list)
    dataframe: pd.DataFrame | None = None
    total_rows: int = 0

    column_types: dict[str, str] = field(default_factory=dict)

    date_format: str | None = None
    decimal_separator: str | None = None
    thousands_separator: str | None = None

    sheet_name: str | None = None
    sheet_names: list[str] = field(default_factory=list)

    warnings: list[str] = field(default_factory=list)
    error: str | None = None


def parse_file(file_bytes: bytes, filename: str) -> ParseResult:
    """Parse a supported file into a normalized dataframe."""

    try:
        if not file_bytes:
            return ParseResult(success=False, filename=filename, error="File is empty")

        extension = Path(filename).suffix.lower()
        if extension == ".xls":
            return ParseResult(
                success=False,
                filename=filename,
                error="Legacy .xls files are not supported. Please re-export as .xlsx or .csv.",
            )

        if extension not in {".csv", ".tsv", ".xlsx"}:
            return ParseResult(
                success=False,
                filename=filename,
                error=f"Unsupported file type: {extension}. Supported: .csv, .tsv, .xlsx",
            )

        if extension in {".csv", ".tsv"}:
            return _parse_csv(file_bytes=file_bytes, filename=filename)

        return _parse_xlsx(file_bytes=file_bytes, filename=filename)
    except Exception as exc:  # pragma: no cover - safety net
        return ParseResult(success=False, filename=filename, error=str(exc))


def _parse_csv(file_bytes: bytes, filename: str) -> ParseResult:
    result = ParseResult(success=False, filename=filename)

    encoding, decoded_text = _detect_encoding(file_bytes)
    result.encoding = encoding

    delimiter = _detect_delimiter(decoded_text)
    result.delimiter = delimiter

    lines = decoded_text.splitlines()
    header_row_index = _detect_header_row_from_lines(lines, delimiter)
    if header_row_index is None:
        header_row_index = 0
        result.warnings.append(
            "No strong header row was found in the first 20 lines; falling back to row 0."
        )
    result.header_row_index = header_row_index

    dataframe = pd.read_csv(
        io.StringIO(decoded_text),
        sep=delimiter,
        dtype=str,
        keep_default_na=False,
        skiprows=header_row_index,
        header=0,
        engine="python",
    )
    dataframe = _clean_dataframe(dataframe)

    result.headers = list(dataframe.columns)
    result.total_rows = len(dataframe)

    (
        dataframe,
        result.column_types,
        result.date_format,
        result.decimal_separator,
        result.thousands_separator,
        result.warnings,
    ) = _detect_and_convert_types(dataframe, result.warnings)

    result.dataframe = dataframe
    result.success = True
    return result


def _parse_xlsx(file_bytes: bytes, filename: str) -> ParseResult:
    result = ParseResult(success=False, filename=filename)

    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    result.sheet_names = list(workbook.sheetnames)

    selected_sheet = _select_sheet_name(workbook)
    result.sheet_name = selected_sheet

    raw_dataframe = pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=selected_sheet,
        dtype=str,
        keep_default_na=False,
        header=None,
        engine="openpyxl",
    )
    header_row_index = _detect_header_row_from_rows(raw_dataframe.head(20).values.tolist())
    if header_row_index is None:
        header_row_index = 0
        result.warnings.append(
            "No strong header row was found in the first 20 rows; falling back to row 0."
        )
    result.header_row_index = header_row_index

    dataframe = _apply_header_row(raw_dataframe, header_row_index)
    dataframe = _clean_dataframe(dataframe)

    result.headers = list(dataframe.columns)
    result.total_rows = len(dataframe)

    (
        dataframe,
        result.column_types,
        result.date_format,
        result.decimal_separator,
        result.thousands_separator,
        result.warnings,
    ) = _detect_and_convert_types(dataframe, result.warnings)

    result.dataframe = dataframe
    result.success = True
    return result


def _detect_encoding(file_bytes: bytes) -> tuple[str, str]:
    detection = chardet.detect(file_bytes)
    detected_encoding = detection.get("encoding")
    confidence = detection.get("confidence") or 0.0

    candidates: list[str] = []
    if detected_encoding and confidence >= 0.7:
        candidates.append(detected_encoding)
    candidates.extend(ENCODING_FALLBACKS)

    seen: set[str] = set()
    for encoding in candidates:
        normalized = encoding.lower()
        if normalized in seen:
            continue
        seen.add(normalized)
        try:
            decoded_text = file_bytes.decode(encoding)
            if any(
                (ord(char) < 32 and char not in "\n\r\t")
                or 127 <= ord(char) <= 159
                or char in "¤©¹»¾"
                for char in decoded_text
            ):
                continue
            return encoding, decoded_text
        except UnicodeDecodeError:
            continue

    raise ValueError("Unable to decode file with supported encodings.")


def _detect_delimiter(decoded_text: str) -> str:
    non_empty_lines = [line for line in decoded_text.splitlines() if line.strip()]
    if not non_empty_lines:
        raise ValueError("File is empty")

    sample = "\n".join(non_empty_lines[:50])[:8192]
    sniffed_delimiter: str | None = None
    try:
        sniffed_delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
    except csv.Error:
        sniffed_delimiter = None

    candidate_lines = non_empty_lines[:5]
    delimiter_scores = {
        delimiter: _score_delimiter(candidate_lines, delimiter)
        for delimiter in [",", ";", "\t"]
    }
    best_delimiter = max(delimiter_scores, key=delimiter_scores.get)

    if sniffed_delimiter is None:
        return best_delimiter

    sniffed_score = delimiter_scores.get(
        sniffed_delimiter,
        (-1, -1, -1, -1, -1),
    )
    if delimiter_scores[best_delimiter] > sniffed_score:
        return best_delimiter
    return sniffed_delimiter


def _score_delimiter(lines: list[str], delimiter: str) -> tuple[int, int, int, int, int]:
    counts = [len(_parse_delimited_line(line, delimiter)) for line in lines if line.strip()]
    if not counts:
        return (-1, -1, -1, -1, -1)

    most_common_count, frequency = Counter(counts).most_common(1)[0]
    all_rows_have_multiple_fields = int(all(count > 1 for count in counts))
    rows_with_multiple_fields = sum(count > 1 for count in counts)
    spread = sum(abs(count - most_common_count) for count in counts)
    unique_counts = len(set(counts))

    return (
        all_rows_have_multiple_fields,
        frequency,
        rows_with_multiple_fields,
        most_common_count,
        -spread - unique_counts,
    )


def _parse_delimited_line(line: str, delimiter: str) -> list[str]:
    return next(csv.reader([line], delimiter=delimiter))


def _detect_header_row_from_lines(lines: list[str], delimiter: str) -> int | None:
    for index, line in enumerate(lines[:20]):
        if not line.strip():
            continue
        score = _score_header_row(_parse_delimited_line(line, delimiter))
        if score >= 4:
            return index
    return None


def _detect_header_row_from_rows(rows: list[list[Any]]) -> int | None:
    for index, row in enumerate(rows[:20]):
        score = _score_header_row(row)
        if score >= 4:
            return index
    return None


def _score_header_row(cells: list[Any]) -> float:
    normalized_cells = [_normalize_header_value(cell) for cell in cells]
    non_empty_cells = [cell for cell in normalized_cells if cell]
    if len(non_empty_cells) < 3:
        return 0.0

    if _looks_like_summary_marker(non_empty_cells[0]):
        return -10.0

    text_like = sum(_looks_text_like(cell) for cell in non_empty_cells)
    value_like = sum(_looks_value_like(cell) for cell in non_empty_cells)
    unique_bonus = 0.5 if len(set(non_empty_cells)) == len(non_empty_cells) else 0.0

    score = len(non_empty_cells) * 0.5 + text_like * 1.5 - value_like * 1.25 + unique_bonus
    if text_like / len(non_empty_cells) >= 0.6:
        score += 2.0
    return score


def _normalize_header_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return _normalize_text(value)
    return _normalize_text(str(value))


def _looks_text_like(value: str) -> bool:
    return any(character.isalpha() for character in value)


def _looks_value_like(value: str) -> bool:
    if not value:
        return False
    if value.startswith("="):
        return True
    if any(pattern.match(value) for pattern in DATE_PATTERNS.values()):
        return True
    if re.fullmatch(r"[-+]?[\d\s,\.]+", value):
        return True
    return False


def _apply_header_row(dataframe: pd.DataFrame, header_row_index: int) -> pd.DataFrame:
    if dataframe.empty:
        return pd.DataFrame()

    normalized = dataframe.copy()
    normalized = normalized.fillna("")
    header_values = [_normalize_header_value(value) for value in normalized.iloc[header_row_index].tolist()]
    data = normalized.iloc[header_row_index + 1 :].reset_index(drop=True).copy()
    data.columns = header_values
    return data


def _clean_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    dataframe = dataframe.copy()
    dataframe = dataframe.map(lambda value: _normalize_text(value) if isinstance(value, str) else value)
    dataframe.columns = [_normalize_header_value(column) for column in dataframe.columns]

    dataframe = dataframe.loc[
        ~dataframe.apply(lambda row: all(_is_blank_value(value) for value in row), axis=1)
    ].reset_index(drop=True)

    footer_index = _find_footer_row_index(dataframe)
    if footer_index is not None:
        dataframe = dataframe.iloc[:footer_index].reset_index(drop=True)

    non_empty_columns = [
        column
        for column in dataframe.columns
        if not all(_is_blank_value(value) for value in dataframe[column].tolist())
    ]
    dataframe = dataframe.loc[:, non_empty_columns]
    return dataframe


def _normalize_text(value: str) -> str:
    return value.replace("\xa0", " ").strip()


def _is_blank_value(value: Any) -> bool:
    return value == "" or value is None


def _find_footer_row_index(dataframe: pd.DataFrame) -> int | None:
    if dataframe.empty or dataframe.shape[1] == 0:
        return None

    first_column = dataframe.columns[0]
    for index, value in enumerate(dataframe[first_column].tolist()):
        text = _normalize_header_value(value)
        if _looks_like_summary_marker(text):
            return index
        if text.startswith("="):
            return index
    return None


def _looks_like_summary_marker(value: str) -> bool:
    lowered = value.casefold()
    return any(marker in lowered for marker in SUMMARY_MARKERS)


def _detect_and_convert_types(
    dataframe: pd.DataFrame,
    warnings: list[str],
) -> tuple[pd.DataFrame, dict[str, str], str | None, str | None, str | None, list[str]]:
    dataframe = dataframe.copy()
    column_types: dict[str, str] = {}
    date_format: str | None = None
    numeric_patterns: list[str] = []

    for column in dataframe.columns:
        original_values = dataframe[column].tolist()
        values = [value if isinstance(value, str) else _normalize_header_value(value) for value in original_values]
        non_empty_values = [value for value in values if not _is_blank_value(value)]

        if not non_empty_values:
            column_types[column] = "empty"
            continue

        detected_date_format = _detect_date_format(non_empty_values)
        if detected_date_format is not None:
            converted_values: list[Any] = []
            for value in values:
                if _is_blank_value(value):
                    converted_values.append("")
                    continue
                try:
                    converted_values.append(_convert_date_value(value, detected_date_format, column, warnings))
                except ValueError:
                    warnings.append(
                        f"Could not convert value {value!r} in column {column!r} to a date."
                    )
                    converted_values.append(value)

            dataframe.loc[:, column] = converted_values
            column_types[column] = "date"
            if date_format is None:
                date_format = detected_date_format
            continue

        numeric_pattern = _detect_numeric_pattern(non_empty_values)
        if numeric_pattern is not None:
            context = _build_numeric_context(non_empty_values)
            converted_values = []
            for value in values:
                if _is_blank_value(value):
                    converted_values.append("")
                    continue
                try:
                    converted_values.append(_convert_numeric_value(value, numeric_pattern, context))
                except ValueError:
                    warnings.append(
                        f"Could not convert value {value!r} in column {column!r} using numeric pattern {numeric_pattern!r}."
                    )
                    converted_values.append(value)

            dataframe.loc[:, column] = converted_values
            column_types[column] = "numeric"
            numeric_patterns.append(numeric_pattern)
            continue

        column_types[column] = "string"

    decimal_separator: str | None = None
    thousands_separator: str | None = None
    if numeric_patterns:
        dominant_pattern = Counter(numeric_patterns).most_common(1)[0][0]
        decimal_separator, thousands_separator = _pattern_separators(dominant_pattern)

    return dataframe, column_types, date_format, decimal_separator, thousands_separator, warnings


def _detect_date_format(values: list[str]) -> str | None:
    pattern_counts = {
        name: sum(bool(regex.match(value)) for value in values)
        for name, regex in DATE_PATTERNS.items()
    }
    best_name, best_count = max(pattern_counts.items(), key=lambda item: item[1])
    if best_count / len(values) <= 0.7:
        return None

    if best_name == "dot":
        return "D.M.YYYY" if any(_has_single_digit_part(value, ".") for value in values) else "DD.MM.YYYY"
    if best_name == "slash":
        return "D/M/YYYY" if any(_has_single_digit_part(value, "/") for value in values) else "DD/MM/YYYY"
    return "YYYY-MM-DD"


def _has_single_digit_part(value: str, separator: str) -> bool:
    parts = value.split(separator)
    return any(len(part) == 1 for part in parts[:2])


def _convert_date_value(
    value: str,
    detected_format: str,
    column: str,
    warnings: list[str],
) -> dt.date:
    if detected_format == "YYYY-MM-DD":
        return dt.date.fromisoformat(value)

    if "/" in value:
        day_text, month_text, year_text = value.split("/")
        day = int(day_text)
        month = int(month_text)
        year = int(year_text)
        if day > 31 or month > 12:
            warnings.append(
                f"Date value {value!r} in column {column!r} falls outside day-first bounds; attempting day-first parsing."
            )
        return dt.date(year, month, day)

    day_text, month_text, year_text = value.split(".")
    return dt.date(int(year_text), int(month_text), int(day_text))


def _detect_numeric_pattern(values: list[str]) -> str | None:
    if not any("." in value or "," in value for value in values):
        return None

    context = _build_numeric_context(values)
    pattern_counts = {
        "space_comma": sum(_matches_numeric_pattern(value, "space_comma", context) for value in values),
        "dot_comma": sum(_matches_numeric_pattern(value, "dot_comma", context) for value in values),
        "comma_dot": sum(_matches_numeric_pattern(value, "comma_dot", context) for value in values),
        "plain_dot": sum(_matches_numeric_pattern(value, "plain_dot", context) for value in values),
    }
    best_pattern, best_count = max(pattern_counts.items(), key=lambda item: item[1])
    if best_count / len(values) <= 0.7:
        return None
    return best_pattern


def _build_numeric_context(values: list[str]) -> dict[str, bool]:
    return {
        "has_dot_group_comma": any(DOT_GROUP_COMMA_RE.match(value) for value in values),
        "has_comma_group_dot": any(COMMA_GROUP_DOT_RE.match(value) for value in values),
    }


def _matches_numeric_pattern(value: str, pattern: str, context: dict[str, bool]) -> bool:
    if pattern == "space_comma":
        return bool(SPACE_GROUP_COMMA_RE.match(value) or SIMPLE_COMMA_DECIMAL_RE.match(value))

    if pattern == "dot_comma":
        if DOT_GROUP_COMMA_RE.match(value):
            return True
        return bool(context["has_dot_group_comma"] and SIMPLE_COMMA_DECIMAL_RE.match(value))

    if pattern == "comma_dot":
        if COMMA_GROUP_DOT_RE.match(value):
            return True
        return bool(context["has_comma_group_dot"] and PLAIN_DOT_DECIMAL_RE.match(value))

    if pattern == "plain_dot":
        return bool(PLAIN_DOT_DECIMAL_RE.match(value))

    return False


def _convert_numeric_value(value: str, pattern: str, context: dict[str, bool]) -> float:
    if not _matches_numeric_pattern(value, pattern, context):
        raise ValueError("Value does not match the selected pattern")

    normalized = value.strip()
    if pattern == "space_comma":
        normalized = normalized.replace(" ", "").replace(",", ".")
    elif pattern == "dot_comma":
        normalized = normalized.replace(".", "").replace(",", ".")
    elif pattern == "comma_dot":
        normalized = normalized.replace(",", "")

    return float(normalized)


def _pattern_separators(pattern: str) -> tuple[str, str | None]:
    if pattern == "space_comma":
        return ",", " "
    if pattern == "dot_comma":
        return ",", "."
    if pattern == "comma_dot":
        return ".", ","
    return ".", None


def _select_sheet_name(workbook: Any) -> str:
    if len(workbook.sheetnames) == 1:
        return workbook.sheetnames[0]

    best_sheet = workbook.sheetnames[0]
    best_score = (-1.0, -1.0, -1.0)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        preview_rows = list(worksheet.iter_rows(min_row=1, max_row=10, values_only=True))
        first_non_empty_row = next(
            (
                [_normalize_header_value(cell) for cell in row]
                for row in preview_rows
                if any(cell not in (None, "") for cell in row)
            ),
            [],
        )
        data_rows = sum(
            1
            for row in worksheet.iter_rows(values_only=True)
            if any(cell not in (None, "") for cell in row)
        )
        header_score = _score_header_row(first_non_empty_row) if first_non_empty_row else 0.0
        score = (
            float(data_rows > 1),
            header_score,
            float(data_rows),
        )
        if score > best_score:
            best_score = score
            best_sheet = sheet_name

    return best_sheet
